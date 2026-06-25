from aiogram import Router
from aiogram.types import (
    Message, BufferedInputFile,
    InlineKeyboardMarkup, InlineKeyboardButton, CallbackQuery,
)
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from datetime import timedelta, timezone, datetime
import asyncio
import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import database as db
import valyuta as val
import charts
from translation import (
    Tkey, say, t, log_exc, GENERIC_ERROR,
    foydalanuvchi_tili, tarjima_qil, register_ui,
)
from .callbacks import CB
from .nav import cb_guard, menu_kb, show, send

# GMT+5 timezone
TOSHKENT_TZ = timezone(timedelta(hours=5))

router = Router()

DAVRLAR = [
    ("bugun", "Bugun"),
    ("kecha", "Kecha"),
    ("hafta", "Shu hafta"),
    ("ohafta", "O'tgan hafta"),
    ("oy", "Shu oy"),
    ("ooy", "O'tgan oy"),
    ("yil", "Shu yil"),
    ("7kun", "Oxirgi 7 kun"),
    ("30kun", "Oxirgi 30 kun"),
    ("custom", "📅 Ixtiyoriy davr"),
]
register_ui(*[label for _, label in DAVRLAR], "📅 Davrni tanlang:", "Hammasi",
            "📦 Mahsulotni tanlang:")


class CustomRange(StatesGroup):
    sana = State()


async def _root_kb(user_id):
    """Inline: hisobot bosh menyusi."""
    return await menu_kb(user_id, [
        [("📊 Hisobot ko'rish", f"{CB.REP_MENU}:korish")],
        [("📁 Fayl yuklash", f"{CB.REP_MENU}:fayl")],
        [("📉 Grafiklar", f"{CB.REP_MENU}:grafik")],
    ])


async def _korish_kb(user_id):
    """Inline: matnli hisobot turlari."""
    return await menu_kb(user_id, [
        [("📊 Umumiy hisobot", f"{CB.REP_MENU}:umumiy"),
         ("📊 Tafsilotli hisobot", f"{CB.REP_MENU}:tafsil")],
        [("💰 Moliya hisoboti", f"{CB.REP_MENU}:moliya"),
         ("👷 Ishchilar hisoboti", f"{CB.REP_MENU}:ishchi")],
        [("🧱 Material sarfi", f"{CB.REP_MENU}:material"),
         ("📈 Taqqoslash", f"{CB.REP_MENU}:taqqos")],
        [("⬅️ Ortga", f"{CB.REP_MENU}:root")],
    ])


async def _fayl_kb(user_id):
    """Inline: eksport formatlari."""
    return await menu_kb(user_id, [
        [("📥 Excel hisobot", f"{CB.REP_MENU}:excel")],
        [("📄 CSV eksport", f"{CB.REP_MENU}:csv"),
         ("📄 PDF eksport", f"{CB.REP_MENU}:pdf")],
        [("⬅️ Ortga", f"{CB.REP_MENU}:root")],
    ])


# ── Davr oralig'i ──
def davr_oraligi(kod):
    bugun = db.bugungi_sana()
    if kod == "bugun":
        return bugun, bugun, "Bugun"
    if kod == "kecha":
        y = bugun - timedelta(days=1)
        return y, y, "Kecha"
    if kod == "hafta":
        start = bugun - timedelta(days=bugun.weekday())
        return start, bugun, "Shu hafta"
    if kod == "ohafta":
        this_mon = bugun - timedelta(days=bugun.weekday())
        last_sun = this_mon - timedelta(days=1)
        last_mon = this_mon - timedelta(days=7)
        return last_mon, last_sun, "O'tgan hafta"
    if kod == "oy":
        return bugun.replace(day=1), bugun, "Shu oy"
    if kod == "ooy":
        first_this = bugun.replace(day=1)
        last_prev = first_this - timedelta(days=1)
        return last_prev.replace(day=1), last_prev, "O'tgan oy"
    if kod == "yil":
        return bugun.replace(month=1, day=1), bugun, "Shu yil"
    if kod == "7kun":
        return bugun - timedelta(days=6), bugun, "Oxirgi 7 kun"
    if kod == "30kun":
        return bugun - timedelta(days=29), bugun, "Oxirgi 30 kun"
    return bugun, bugun, "Bugun"


def oldingi_davr(boshliq, oxiri):
    uzunlik = (oxiri - boshliq).days + 1
    prev_oxiri = boshliq - timedelta(days=1)
    prev_boshliq = prev_oxiri - timedelta(days=uzunlik - 1)
    return prev_boshliq, prev_oxiri


async def davr_keyboard(user_id, rtype, pid_str):
    til = await foydalanuvchi_tili(user_id)
    kb, row = [], []
    for kod, label in DAVRLAR:
        matn = label if til == "uz" else await tarjima_qil(label, til)
        row.append(InlineKeyboardButton(
            text=matn, callback_data=f"{CB.REP}:{rtype}:{pid_str}:{kod}"))
        if len(row) == 2:
            kb.append(row)
            row = []
    if row:
        kb.append(row)
    orqaga = "⬅️ Ortga" if til == "uz" else await tarjima_qil("⬅️ Ortga", til)
    kb.append([InlineKeyboardButton(text=orqaga, callback_data=f"{CB.REP_MENU}:root")])
    return InlineKeyboardMarkup(inline_keyboard=kb)


async def product_filter_keyboard(user_id, rtype):
    til = await foydalanuvchi_tili(user_id)
    hammasi = "Hammasi" if til == "uz" else await tarjima_qil("Hammasi", til)
    kb = [[InlineKeyboardButton(text=f"🌐 {hammasi}", callback_data=f"{CB.REP_PROD}:{rtype}:all")]]
    for p in await db.get_mahsulotlar(faqat_faol=False):
        kb.append([InlineKeyboardButton(
            text=f"{p['emoji']} {p['nomi']}", callback_data=f"{CB.REP_PROD}:{rtype}:{p['id']}")])
    orqaga = "⬅️ Ortga" if til == "uz" else await tarjima_qil("⬅️ Ortga", til)
    kb.append([InlineKeyboardButton(text=orqaga, callback_data=f"{CB.REP_MENU}:root")])
    return InlineKeyboardMarkup(inline_keyboard=kb)


# ── Yordamchilar ──
def _vaqt_str(vaqt, fmt="%d.%m %H:%M"):
    if hasattr(vaqt, "strftime"):
        if vaqt.tzinfo is None:
            vaqt = vaqt.replace(tzinfo=timezone.utc).astimezone(TOSHKENT_TZ)
        return vaqt.strftime(fmt)
    return str(vaqt)[:16]


def _delta(cur, prev):
    if prev == 0:
        return "(▲ yangi)" if cur > 0 else ""
    p = (cur - prev) / prev * 100
    arrow = "▲" if p >= 0 else "▼"
    return f"({arrow} {abs(p):.1f}%)"


def _group_by_product(rows):
    out = {}
    for r in rows:
        d = out.setdefault(r["product_id"], {"nomi": r["product_nomi"], "items": []})
        d["items"].append(r)
    return out


async def _narx_maps():
    tannarx = await db.get_block_tannarx_map()
    sotuv = {}
    for m in await db.get_mahsulotlar(faqat_faol=False):
        for b in await db.get_bloklar(m["id"]):
            sotuv[(m["id"], b["kod"])] = b["sotuv_narx"] or 0
    return tannarx, sotuv


async def _label_pid(product_id):
    if not product_id:
        return "Hammasi"
    p = await db.get_mahsulot(product_id)
    return p["nomi"] if p else "?"


async def ombor_holati():
    try:
        materials = await db.get_materials()
        if not materials:
            return "   Ma'lumot yo'q\n"
        text = ""
        for m in materials:
            qoldiq_asl = db.asosiydan_birlikga(m[2], m[4])
            text += f"   {m[1]}: {qoldiq_asl:.2f} {m[4]}\n"
        return text
    except Exception:
        return "   Xatolik\n"


async def tayyor_holati(product_id=None):
    try:
        goods = await db.get_all_finished_goods()
        if product_id:
            goods = [g for g in goods if g["product_id"] == product_id]
        if not goods:
            return "   Ma'lumot yo'q\n"
        text = ""
        jami = 0
        joriy = None
        for g in goods:
            if g["product_id"] != joriy:
                joriy = g["product_id"]
                text += f"   {g['emoji']} {g['product_nomi']}:\n"
            text += f"      {g['nomi']}: {g['qoldiq']} ta\n"
            jami += g["qoldiq"]
        text += f"   Jami: {jami} ta\n"
        return text
    except Exception:
        return "   Xatolik\n"


# ── Hisobot matnlari ──
async def hisobot_matni(boshliq, oxiri, sarlavha, product_id=None):
    try:
        qol, pblk, sblk, ombor, tayyor = await asyncio.gather(
            db.get_production_qolip_range(boshliq, oxiri, product_id),
            db.get_production_blocks_range(boshliq, oxiri, product_id),
            db.get_sales_blocks_range(boshliq, oxiri, product_id),
            ombor_holati(),
            tayyor_holati(product_id),
        )

        text = (f"{sarlavha}\n📅 {boshliq} — {oxiri}\n"
                f"━━━━━━━━━━━━━━━━\n\n🏭 Ishlab chiqarish:\n")
        if qol:
            pgr = _group_by_product(pblk)
            for pid, info in qol.items():
                text += f"   {info['nomi']}: {info['qolip']} qolip\n"
                for it in pgr.get(pid, {}).get("items", []):
                    text += f"      {it['blok_nomi']}: {it['soni']} ta\n"
        else:
            text += "   Ma'lumot yo'q\n"

        text += "\n💰 Sotuv:\n"
        if sblk:
            for pid, info in _group_by_product(sblk).items():
                text += f"   {info['nomi']}:\n"
                jami = 0
                for it in info["items"]:
                    text += f"      {it['blok_nomi']}: {it['qty']} ta\n"
                    jami += it["qty"]
                text += f"      Jami: {jami} ta\n"
        else:
            text += "   Ma'lumot yo'q\n"

        text += f"\n🏬 Tayyor mahsulot:\n{tayyor}\n"
        text += f"🏪 Xom ashyo qoldig'i:\n{ombor}"
        return text
    except Exception as e:
        log_exc("hisobot_matni", e)
        return GENERIC_ERROR


async def gen_tafsil(boshliq, oxiri, sarlavha, product_id=None):
    prod_detail = await db.get_production_detail_range(boshliq, oxiri, product_id)
    sales_detail = await db.get_sales_detail_range(boshliq, oxiri, product_id)
    chiqim = await db.get_material_sarfi(boshliq, oxiri, product_id)

    text = f"📊 Tafsilotli hisobot ({sarlavha})\n📅 {boshliq} — {oxiri}\n━━━━━━━━━━━━━━━━\n\n"
    text += "🏭 ISHLAB CHIQARISH:\n"
    if prod_detail:
        for p in prod_detail[:15]:
            text += (f"   {_vaqt_str(p['vaqt'])} | {p.get('user_ism') or 'Nomalum'}\n"
                     f"   {p.get('product_nomi') or ''} · {p.get('shablon_nomi') or '?'}: "
                     f"{p['qolip_soni']} qolip\n")
    else:
        text += "   Ma'lumot yo'q\n"

    text += "\n💰 SOTUV:\n"
    if sales_detail:
        for s in sales_detail[:15]:
            text += (f"   {_vaqt_str(s['vaqt'])} | {s.get('user_ism') or 'Nomalum'}\n"
                     f"   {s.get('product_nomi') or ''} · "
                     f"{s.get('blok_nomi') or s['block_type']}: {s['miqdor']} ta\n")
    else:
        text += "   Ma'lumot yo'q\n"

    text += "\n📉 XOM ASHYO SARFI:\n"
    if chiqim:
        for ch in chiqim:
            asl = db.asosiydan_birlikga(ch["jami"], ch["birlik"])
            text += f"   {ch['nomi']}: {asl:.2f} {ch['birlik']}\n"
    else:
        text += "   Ma'lumot yo'q\n"
    return text


async def gen_moliya(boshliq, oxiri, sarlavha, product_id=None):
    sblk, narx, xom, goods = await asyncio.gather(
        db.get_sales_blocks_range(boshliq, oxiri, product_id),
        _narx_maps(),
        db.ombor_xom_qiymati(),
        db.get_all_finished_goods(),
    )
    tannarx, sotuv = narx
    daromad = sum(x["rev"] for x in sblk)
    cogs = sum(x["qty"] * tannarx.get((x["product_id"], x["kod"]), 0) for x in sblk)
    foyda = daromad - cogs
    foyda_foiz = (foyda / daromad * 100) if daromad > 0 else 0.0

    if product_id:
        goods = [g for g in goods if g["product_id"] == product_id]
    tayyor_tannarx = sum(g["qoldiq"] * tannarx.get((g["product_id"], g["kod"]), 0) for g in goods)
    tayyor_sotuv = sum(g["qoldiq"] * sotuv.get((g["product_id"], g["kod"]), 0) for g in goods)
    kod = await val.get_active()

    text = (
        f"💰 Moliya hisoboti ({sarlavha})\n"
        f"📅 {boshliq} — {oxiri}\n"
        f"💱 Valyuta: {val.belgi(kod)} ({kod})\n"
        f"━━━━━━━━━━━━━━━━\n\n📈 SOTUV:\n")
    for pid, info in _group_by_product(sblk).items():
        text += f"   {info['nomi']}:\n"
        for it in info["items"]:
            text += f"      {it['blok_nomi']}: {it['qty']} ta = {await val.format_uzs(it['rev'])}\n"
    text += (
        f"\n   Daromad: {await val.format_uzs(daromad)}\n"
        f"📉 Tannarx (COGS): {await val.format_uzs(cogs)}\n"
        f"💵 Sof foyda: {await val.format_uzs(foyda)} ({foyda_foiz:.1f}%)\n\n"
        f"🏪 Ombor qiymati (xom ashyo): {await val.format_uzs(xom)}\n"
        f"🏬 Tayyor mahsulot:\n"
        f"   Tannarxda: {await val.format_uzs(tayyor_tannarx)}\n"
        f"   Sotuv narxida: {await val.format_uzs(tayyor_sotuv)}")
    return text


async def gen_ishchi(boshliq, oxiri, sarlavha, product_id=None):
    prod = await db.get_production_by_user_range(boshliq, oxiri)
    sales = await db.get_sales_by_user_range(boshliq, oxiri)

    text = (f"👷 Ishchilar hisoboti ({sarlavha})\n📅 {boshliq} — {oxiri}\n"
            f"━━━━━━━━━━━━━━━━\n\n🏭 ISHLAB CHIQARISH (ishbay haq):\n")
    if prod:
        for p in prod:
            text += (f"   {p['ism']}: {p['qolip']} qolip\n"
                     f"      💵 Ish haqi: {await val.format_uzs(p['haq'])}\n")
    else:
        text += "   Ma'lumot yo'q\n"

    text += "\n💰 SOTUV (kim sotdi):\n"
    if sales:
        for s in sales:
            text += f"   {s['ism']}: {s['qty']} ta = {await val.format_uzs(s['rev'])}\n"
    else:
        text += "   Ma'lumot yo'q\n"
    return text


async def _metrikalar(boshliq, oxiri, product_id=None):
    qol, sblk, narx = await asyncio.gather(
        db.get_production_qolip_range(boshliq, oxiri, product_id),
        db.get_sales_blocks_range(boshliq, oxiri, product_id),
        _narx_maps(),
    )
    tannarx, _ = narx
    jami_qolip = sum(v["qolip"] for v in qol.values())
    daromad = sum(x["rev"] for x in sblk)
    sotildi = sum(x["qty"] for x in sblk)
    cogs = sum(x["qty"] * tannarx.get((x["product_id"], x["kod"]), 0) for x in sblk)
    return {"qolip": jami_qolip, "sotuv": sotildi,
            "daromad": daromad, "foyda": daromad - cogs}


async def gen_taqqos(boshliq, oxiri, sarlavha, product_id=None):
    pb, po = oldingi_davr(boshliq, oxiri)
    cur = await _metrikalar(boshliq, oxiri, product_id)
    prev = await _metrikalar(pb, po, product_id)
    return (
        f"📈 Taqqoslash: {sarlavha}\n📅 {boshliq} — {oxiri}\n"
        f"   (oldingi davr: {pb} — {po})\n━━━━━━━━━━━━━━━━\n\n"
        f"🏭 Jami qolip: {cur['qolip']} {_delta(cur['qolip'], prev['qolip'])}\n"
        f"💰 Sotuv (dona): {cur['sotuv']} {_delta(cur['sotuv'], prev['sotuv'])}\n"
        f"   Daromad: {await val.format_uzs(cur['daromad'])} "
        f"{_delta(cur['daromad'], prev['daromad'])}\n"
        f"   Foyda: {await val.format_uzs(cur['foyda'])} "
        f"{_delta(cur['foyda'], prev['foyda'])}")


async def gen_material(boshliq, oxiri, sarlavha, product_id=None):
    sarf = await db.get_material_sarfi(boshliq, oxiri, product_id)
    if not sarf:
        return (f"🧱 Material sarfi ({sarlavha})\n📅 {boshliq} — {oxiri}\n\n   Ma'lumot yo'q")
    text = (f"🧱 Material sarfi ({sarlavha})\n📅 {boshliq} — {oxiri}\n━━━━━━━━━━━━━━━━\n\n")
    jami_narx = 0.0
    for s in sarf:
        disp = db.asosiydan_birlikga(s["jami"], s["birlik"])
        narx_summa = s["jami"] * s["narx"]
        jami_narx += narx_summa
        text += (f"   {s['nomi']}: {disp:.2f} {s['birlik']} "
                 f"= {await val.format_uzs(narx_summa)}\n")
    text += f"\n💰 Jami sarf qiymati: {await val.format_uzs(jami_narx)}"
    return text


# ── Excel ──
async def _excel_yubor(target, user_id, boshliq, oxiri, product_id=None):
    pblk, qol, sblk, materials, goods, audit_logs, chiqim, narx = await asyncio.gather(
        db.get_production_blocks_range(boshliq, oxiri, product_id),
        db.get_production_qolip_range(boshliq, oxiri, product_id),
        db.get_sales_blocks_range(boshliq, oxiri, product_id),
        db.get_materials(),
        db.get_all_finished_goods(),
        db.get_audit_log(200),
        db.get_material_sarfi(boshliq, oxiri, product_id),
        _narx_maps(),
    )
    tannarx, sotuv = narx
    if product_id:
        goods = [g for g in goods if g["product_id"] == product_id]

    wb = openpyxl.Workbook()
    sarlavha_font = Font(bold=True, size=11, color="FFFFFF")
    sarlavha_fill = PatternFill("solid", fgColor="2E75B6")
    markaz = Alignment(horizontal="center", vertical="center")

    def hdr(ws, qator, ustunlar):
        for col, txt in enumerate(ustunlar, 1):
            cell = ws.cell(row=qator, column=col, value=txt)
            cell.font = sarlavha_font
            cell.fill = sarlavha_fill
            cell.alignment = markaz

    ws1 = wb.active
    ws1.title = "Ishlab chiqarish"
    hdr(ws1, 1, ["Mahsulot", "Blok", "Ishlab chiqarildi (dona)"])
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 22
    for x in pblk:
        ws1.append([x["product_nomi"], x["blok_nomi"], x["soni"]])
    ws1.append([])
    for pid, info in qol.items():
        ws1.append([info["nomi"], "JAMI QOLIP", info["qolip"]])

    ws2 = wb.create_sheet("Sotuv")
    hdr(ws2, 1, ["Mahsulot", "Blok", "Sotildi (dona)", "Daromad (so'm)"])
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 22
    for x in sblk:
        ws2.append([x["product_nomi"], x["blok_nomi"], x["qty"], round(x["rev"])])

    ws3 = wb.create_sheet("Xom ashyo sarfi")
    hdr(ws3, 1, ["Material", "Ketgan miqdor", "Birlik", "Qiymat (so'm)"])
    ws3.column_dimensions["A"].width = 20
    for ch in chiqim:
        asl = db.asosiydan_birlikga(ch["jami"], ch["birlik"])
        ws3.append([ch["nomi"], round(asl, 2), ch["birlik"], round(ch["jami"] * ch["narx"])])

    ws4 = wb.create_sheet("Ombor qoldiqlari")
    hdr(ws4, 1, ["Material", "Qoldiq", "Birlik"])
    ws4.column_dimensions["A"].width = 20
    for m in materials:
        ws4.append([m[1], round(db.asosiydan_birlikga(m[2], m[4]), 2), m[4]])
    ws4.append([])
    ws4.append(["── Tayyor mahsulot ──", "", ""])
    for g in goods:
        ws4.append([f"{g['product_nomi']} · {g['nomi']}", g["qoldiq"], "ta"])

    ws5 = wb.create_sheet("Moliya")
    hdr(ws5, 1, ["Ko'rsatkich", "Qiymat (so'm)"])
    ws5.column_dimensions["A"].width = 30
    ws5.column_dimensions["B"].width = 20
    daromad = sum(x["rev"] for x in sblk)
    cogs = sum(x["qty"] * tannarx.get((x["product_id"], x["kod"]), 0) for x in sblk)
    xom = await db.ombor_xom_qiymati()
    tayyor_tannarx = sum(g["qoldiq"] * tannarx.get((g["product_id"], g["kod"]), 0) for g in goods)
    tayyor_sotuv = sum(g["qoldiq"] * sotuv.get((g["product_id"], g["kod"]), 0) for g in goods)
    for nomi_q, qiymat in [
        ("Jami daromad", round(daromad)), ("Tannarx (COGS)", round(cogs)),
        ("Sof foyda", round(daromad - cogs)),
        ("Ombor (xom ashyo) qiymati", round(xom)),
        ("Tayyor mahsulot (tannarx)", round(tayyor_tannarx)),
        ("Tayyor mahsulot (sotuv)", round(tayyor_sotuv)),
    ]:
        ws5.append([nomi_q, qiymat])

    ws6 = wb.create_sheet("Audit log")
    hdr(ws6, 1, ["Vaqt", "Foydalanuvchi", "Rol", "Amal", "Tafsilot"])
    ws6.column_dimensions["A"].width = 18
    ws6.column_dimensions["B"].width = 16
    ws6.column_dimensions["D"].width = 25
    ws6.column_dimensions["E"].width = 40
    for log in audit_logs:
        ws6.append([_vaqt_str(log["vaqt"], "%d.%m.%Y %H:%M"), log["ism"] or "",
                    log["rol"] or "", log["amal"] or "", log["tafsilot"] or ""])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    fayl_nomi = f"hisobot_{boshliq}_{oxiri}.xlsx"
    caption = await t(f"📥 Excel hisobot\n📅 {boshliq} — {oxiri}", user_id)
    await target.answer_document(
        BufferedInputFile(buffer.read(), filename=fayl_nomi), caption=caption)


# ── Grafiklar ──
async def _grafik_yubor(target, user_id, boshliq, oxiri, product_id=None):
    if not charts.MATPLOTLIB:
        await _yubor(target, user_id, "❌ Grafik moduli mavjud emas (matplotlib o'rnatilmagan).")
        return
    prod_daily = await db.get_production_daily(boshliq, oxiri, product_id)
    sales_daily = await db.get_sales_daily(boshliq, oxiri, product_id)
    if not prod_daily and not sales_daily:
        await _yubor(target, user_id, "📊 Bu davrda ma'lumot yo'q.")
        return

    sanalar = sorted(set([d["sana"] for d in prod_daily] + [d["sana"] for d in sales_daily]))
    prod_map = {d["sana"]: d for d in prod_daily}
    sales_map = {d["sana"]: d for d in sales_daily}
    labels = [str(s)[5:] for s in sanalar]
    qolip_vals = [prod_map.get(s, {}).get("qolip", 0) for s in sanalar]
    sotuv_vals = [sales_map.get(s, {}).get("qty", 0) for s in sanalar]

    png = charts.trend_chart(labels, qolip_vals, sotuv_vals, "Ishlab chiqarish va sotuv")
    if png:
        cap = await t("📈 Ishlab chiqarish va sotuv trendi", user_id)
        await target.answer_photo(BufferedInputFile(png, "trend.png"), caption=cap)

    # Kunlik daromad va (taqribiy) foyda
    sblk = await db.get_sales_blocks_range(boshliq, oxiri, product_id)
    tannarx, _ = await _narx_maps()
    daromad_jami = sum(x["rev"] for x in sblk)
    cogs_jami = sum(x["qty"] * tannarx.get((x["product_id"], x["kod"]), 0) for x in sblk)
    margin = ((daromad_jami - cogs_jami) / daromad_jami) if daromad_jami > 0 else 0.0

    kod = await val.get_active()
    rate = await val.get_rate(kod)
    conv = (lambda x: x) if (kod == val.ASOS or not rate) else (lambda x: x / rate)
    rev_vals, profit_vals = [], []
    for s in sanalar:
        rev = sales_map.get(s, {}).get("rev", 0)
        rev_vals.append(round(conv(rev), 2))
        profit_vals.append(round(conv(rev * margin), 2))
    if any(rev_vals):
        png3 = charts.finance_bar(labels, rev_vals, profit_vals, val.belgi(kod))
        if png3:
            cap = await t("📊 Kunlik daromad va (taqribiy) foyda", user_id)
            await target.answer_photo(BufferedInputFile(png3, "finance.png"), caption=cap)


# ── CSV / PDF ──
async def _csv_yubor(target, user_id, boshliq, oxiri, product_id=None):
    prod = await db.get_production_daily(boshliq, oxiri, product_id)
    sales = await db.get_sales_daily(boshliq, oxiri, product_id)
    pmap = {d["sana"]: d for d in prod}
    smap = {d["sana"]: d for d in sales}
    sanalar = sorted(set(list(pmap.keys()) + list(smap.keys())))
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["Sana", "Qolip", "Sotuv (dona)", "Daromad (so'm)"])
    for s in sanalar:
        p = pmap.get(s, {})
        sd = smap.get(s, {})
        w.writerow([s, p.get("qolip", 0), sd.get("qty", 0), round(sd.get("rev", 0))])
    data = ("\ufeff" + out.getvalue()).encode("utf-8")
    fn = f"hisobot_{boshliq}_{oxiri}.csv"
    cap = await t("📄 CSV eksport", user_id)
    await target.answer_document(BufferedInputFile(data, fn), caption=cap)


async def _pdf_yubor(target, user_id, boshliq, oxiri, product_id=None):
    if not charts.MATPLOTLIB:
        await _yubor(target, user_id, "❌ PDF moduli mavjud emas (matplotlib o'rnatilmagan).")
        return
    umumiy = await hisobot_matni(boshliq, oxiri, "Umumiy hisobot", product_id)
    moliya = await gen_moliya(boshliq, oxiri, "Moliya", product_id)
    body = await t(umumiy + "\n\n" + moliya, user_id)
    png = charts.text_pdf("Hisobot", body)
    if not png:
        await _yubor(target, user_id, "❌ PDF yaratilmadi.")
        return
    fn = f"hisobot_{boshliq}_{oxiri}.pdf"
    cap = await t("📄 PDF eksport", user_id)
    await target.answer_document(BufferedInputFile(png, fn), caption=cap)


# ── Umumiy yuborish ──
async def _yubor(target, user_id, matn):
    tt = await t(matn, user_id)
    if len(tt) > 4096:
        tt = tt[:4095] + "…"
    await target.answer(tt)


async def _generate(target, user_id, rtype, boshliq, oxiri, sarlavha, product_id=None):
    try:
        if product_id:
            sarlavha = f"{await _label_pid(product_id)} · {sarlavha}"
        if rtype == "excel":
            await _excel_yubor(target, user_id, boshliq, oxiri, product_id)
            return
        if rtype == "grafik":
            await _grafik_yubor(target, user_id, boshliq, oxiri, product_id)
            return
        if rtype == "csv":
            await _csv_yubor(target, user_id, boshliq, oxiri, product_id)
            return
        if rtype == "pdf":
            await _pdf_yubor(target, user_id, boshliq, oxiri, product_id)
            return
        if rtype == "umumiy":
            matn = await hisobot_matni(boshliq, oxiri, f"📊 {sarlavha}", product_id)
        elif rtype == "tafsil":
            matn = await gen_tafsil(boshliq, oxiri, sarlavha, product_id)
        elif rtype == "moliya":
            matn = await gen_moliya(boshliq, oxiri, sarlavha, product_id)
        elif rtype == "ishchi":
            matn = await gen_ishchi(boshliq, oxiri, sarlavha, product_id)
        elif rtype == "material":
            matn = await gen_material(boshliq, oxiri, sarlavha, product_id)
        elif rtype == "taqqos":
            matn = await gen_taqqos(boshliq, oxiri, sarlavha, product_id)
        else:
            matn = await hisobot_matni(boshliq, oxiri, sarlavha, product_id)
        await _yubor(target, user_id, matn)
    except Exception as e:
        log_exc("report_generate", e)
        await _yubor(target, user_id, GENERIC_ERROR)


async def _ruxsat(user_id, rtype):
    user = await db.get_user(user_id)
    if not user or not user["faol"]:
        return False
    if user["rol"] == "superadmin":
        return True
    perm = "excel_hisobot" if rtype in ("excel", "csv", "pdf") else (
        "moliya_korish" if rtype in ("moliya", "ishchi") else "hisobot_korish")
    return await db.has_permission(user_id, user["rol"], perm)


# ── Kirish ──
@router.message(Tkey("📊 Hisobot"))
async def hisobot(message: Message):
    await send(message, "📊 Hisobotlar:", await _root_kb(message.from_user.id))


async def _davr_sorov_cb(callback, rtype):
    uid = callback.from_user.id
    prods = await db.get_mahsulotlar(faqat_faol=False)
    if len(prods) <= 1:
        await show(callback, "📅 Davrni tanlang:", await davr_keyboard(uid, rtype, "all"))
    else:
        await show(callback, "📦 Mahsulotni tanlang:",
                   await product_filter_keyboard(uid, rtype))


@router.callback_query(lambda c: c.data and c.data.startswith(f"{CB.REP_MENU}:"))
async def repmenu_cb(callback: CallbackQuery):
    arg = callback.data.split(":", 1)[1]
    uid = callback.from_user.id
    if arg in ("root", "korish", "fayl"):
        if not await cb_guard(callback):
            return
        if arg == "root":
            await show(callback, "📊 Hisobotlar:", await _root_kb(uid))
        elif arg == "korish":
            await show(callback, "📊 Hisobot turini tanlang:", await _korish_kb(uid))
        else:
            await show(callback, "📁 Yuklab olish formatini tanlang:", await _fayl_kb(uid))
        await callback.answer()
        return
    # Leaf: hisobot turi — ruxsat tekshiriladi, so'ng davr/mahsulot tanlash
    if not await _ruxsat(uid, arg):
        await callback.answer("⛔ Ruxsat yo'q!", show_alert=True)
        return
    await _davr_sorov_cb(callback, arg)
    await callback.answer()


# ── Mahsulot tanlash callback ──
@router.callback_query(lambda c: c.data and c.data.startswith(f"{CB.REP_PROD}:"))
async def repp_callback(callback: CallbackQuery):
    parts = callback.data.split(":")
    rtype = parts[1] if len(parts) > 1 else "umumiy"
    pid_str = parts[2] if len(parts) > 2 else "all"
    if not await _ruxsat(callback.from_user.id, rtype):
        await callback.answer("⛔", show_alert=False)
        return
    await callback.answer()
    await callback.message.edit_text(
        await t("📅 Davrni tanlang:", callback.from_user.id),
        reply_markup=await davr_keyboard(callback.from_user.id, rtype, pid_str))


# ── Davr callback ──
@router.callback_query(lambda c: c.data and c.data.startswith(f"{CB.REP}:"))
async def rep_callback(callback: CallbackQuery, state: FSMContext):
    parts = callback.data.split(":")
    rtype = parts[1] if len(parts) > 1 else "umumiy"
    pid_str = parts[2] if len(parts) > 2 else "all"
    period = parts[3] if len(parts) > 3 else "bugun"
    product_id = None if pid_str == "all" else int(pid_str)

    if not await _ruxsat(callback.from_user.id, rtype):
        await callback.answer("⛔", show_alert=False)
        return
    await callback.answer()

    if period == "custom":
        await state.clear()
        await state.update_data(rtype=rtype, product_id=product_id)
        await state.set_state(CustomRange.sana)
        xabar = await t(
            "📅 Davrni kiriting (YYYY-MM-DD YYYY-MM-DD):\nMisol: 2026-06-01 2026-06-15",
            callback.from_user.id)
        await callback.message.answer(xabar)
        return

    boshliq, oxiri, sarlavha = davr_oraligi(period)
    await _generate(callback.message, callback.from_user.id, rtype,
                    boshliq, oxiri, sarlavha, product_id)


@router.message(CustomRange.sana)
async def custom_sana(message: Message, state: FSMContext):
    try:
        parts = message.text.strip().split()
        if len(parts) != 2:
            raise ValueError
        b = datetime.strptime(parts[0], "%Y-%m-%d").date()
        o = datetime.strptime(parts[1], "%Y-%m-%d").date()
        if b > o:
            b, o = o, b
        data = await state.get_data()
        rtype = data.get("rtype", "umumiy")
        product_id = data.get("product_id")
        await state.clear()
        await _generate(message, message.from_user.id, rtype, b, o, f"{b} — {o}", product_id)
    except ValueError:
        await say(message, "❌ Format: YYYY-MM-DD YYYY-MM-DD\nMisol: 2026-06-01 2026-06-15")


async def avtomatik_hisobot(bot, chat_id, turi="kun"):
    try:
        if turi == "hafta":
            b, o, _ = davr_oraligi("ohafta")
            sarlavha = "🔔 Avtomatik haftalik hisobot"
        elif turi == "oy":
            b, o, _ = davr_oraligi("ooy")
            sarlavha = "🔔 Avtomatik oylik hisobot"
        else:
            bugun = db.bugungi_sana()
            b = o = bugun
            sarlavha = "🔔 Avtomatik kunlik hisobot"
        text = await hisobot_matni(b, o, sarlavha, None)
        text = await t(text, chat_id)
        await bot.send_message(chat_id, text)
    except Exception as e:
        log_exc("avtomatik_hisobot", e)
