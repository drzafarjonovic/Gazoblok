"""Gazoblok ERP — Web/Mobil API (FastAPI).

Ishga tushirish:
    uvicorn webapi.main:app --host 0.0.0.0 --port 8000

Bot bilan bir xil bazaga ulanadi va mavjud `database` funksiyalarini qayta
ishlatadi. Bot kodiga o'zgartirish kiritilmagan.
"""
import os
import io
import time

from fastapi import FastAPI, Depends, HTTPException, Header, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel

import database as db
from database.core import (
    get_pool, birlikni_asosiyga, asosiydan_birlikga, bugungi_sana,
)
from . import auth

app = FastAPI(title="Gazoblok ERP API", version="0.5.0")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


async def _ensure_auth_columns():
    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS telefon TEXT")
        await conn.execute("ALTER TABLE users ADD COLUMN IF NOT EXISTS parol_hash TEXT")
        await conn.execute(
            "CREATE UNIQUE INDEX IF NOT EXISTS idx_users_telefon "
            "ON users(telefon) WHERE telefon IS NOT NULL"
        )


@app.on_event("startup")
async def _startup():
    await _ensure_auth_columns()


# ═══ AUTENTIFIKATSIYA ═══
class LoginIn(BaseModel):
    telefon: str
    parol: str


async def current_user(authorization: str = Header(default="")):
    if not authorization.startswith("Bearer "):
        raise HTTPException(401, "Token yo'q")
    token = authorization.split(" ", 1)[1]
    try:
        payload = auth.decode_token(token)
    except Exception:
        raise HTTPException(401, "Token yaroqsiz yoki muddati tugagan")
    user = await db.get_user(int(payload["sub"]))
    if not user:
        raise HTTPException(401, "Foydalanuvchi topilmadi yoki bloklangan")
    return user


async def _perms(user) -> dict:
    return await db.get_user_permissions(user["id"], user["rol"])


async def require_perm(user, kalit: str) -> dict:
    perms = await _perms(user)
    if not perms.get(kalit):
        raise HTTPException(403, "Sizda bu amal uchun ruxsat yo'q")
    return perms


@app.post("/api/auth/login")
async def login(data: LoginIn):
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, ism, rol, faol, parol_hash FROM users WHERE telefon=$1",
            data.telefon.strip(),
        )
    if not row or not row["parol_hash"]:
        raise HTTPException(401, "Telefon yoki parol noto'g'ri")
    if not row["faol"]:
        raise HTTPException(403, "Hisob bloklangan")
    if not auth.verify_password(data.parol, row["parol_hash"]):
        raise HTTPException(401, "Telefon yoki parol noto'g'ri")
    return {
        "token": auth.create_token(row["id"], row["rol"]),
        "user": {"id": row["id"], "ism": row["ism"], "rol": row["rol"]},
    }


class TgLoginIn(BaseModel):
    init_data: str


def _validate_tg_init_data(init_data: str, bot_token: str, max_age: int = 86400):
    import hmac as _hmac, hashlib as _hl, json as _json
    from urllib.parse import parse_qsl
    try:
        parsed = dict(parse_qsl(init_data, keep_blank_values=True))
    except Exception:
        return None
    recv = parsed.pop("hash", None)
    if not recv:
        return None
    dcs = "\n".join(f"{k}={parsed[k]}" for k in sorted(parsed))
    secret = _hmac.new(b"WebAppData", bot_token.encode(), _hl.sha256).digest()
    calc = _hmac.new(secret, dcs.encode(), _hl.sha256).hexdigest()
    if not _hmac.compare_digest(calc, recv):
        return None
    try:
        ad = int(parsed.get("auth_date", "0"))
        if max_age and ad and (int(time.time()) - ad) > max_age:
            return None
    except Exception:
        pass
    try:
        return _json.loads(parsed.get("user", "{}"))
    except Exception:
        return None


@app.post("/api/auth/telegram")
async def telegram_login(data: TgLoginIn):
    bot_token = os.getenv("BOT_TOKEN")
    if not bot_token:
        raise HTTPException(500, "BOT_TOKEN sozlanmagan")
    tg_user = _validate_tg_init_data(data.init_data, bot_token)
    if not tg_user or not tg_user.get("id"):
        raise HTTPException(401, "Telegram imzosi yaroqsiz")
    tg_id = int(tg_user["id"])
    pool = await get_pool()
    async with pool.acquire() as conn:
        row = await conn.fetchrow(
            "SELECT id, ism, rol, faol FROM users WHERE id=$1", tg_id
        )
    if not row:
        raise HTTPException(403, "Bu Telegram hisobi tizimda yo'q")
    if not row["faol"]:
        raise HTTPException(403, "Hisob bloklangan")
    return {
        "token": auth.create_token(row["id"], row["rol"]),
        "user": {"id": row["id"], "ism": row["ism"], "rol": row["rol"]},
    }


@app.get("/api/me")
async def me(user=Depends(current_user)):
    perms = await _perms(user)
    return {"id": user["id"], "ism": user.get("ism"), "rol": user.get("rol"),
            "permissions": perms}


# ═══ O'QISH: mahsulot / material / tayyor ═══
@app.get("/api/products")
async def products(user=Depends(current_user)):
    return await db.get_mahsulotlar()


async def _materials_enriched():
    """Materiallar + min chegara + kam qolgan bayrog'i (asl birlikda)."""
    rows = await db.get_materials()  # (id, nomi, qoldiq_baza, birlik_baza, asl_birlik)
    settings = await db.get_settings()  # (nomi, min_chegara_baza, birlik_baza, id, asl_birlik)
    min_map = {s[3]: s[1] for s in settings}
    natija = []
    for r in rows:
        mid, nomi, qoldiq_baza, _bir, asl = r[0], r[1], r[2], r[3], r[4]
        min_baza = min_map.get(mid)
        natija.append({
            "id": mid, "nomi": nomi,
            "qoldiq": round(asosiydan_birlikga(qoldiq_baza, asl), 2),
            "birlik": asl,
            "qoldiq_baza": qoldiq_baza,
            "min": round(asosiydan_birlikga(min_baza, asl), 2) if min_baza else None,
            "kam": bool(min_baza) and qoldiq_baza <= min_baza,
        })
    return natija


@app.get("/api/materials")
async def materials(user=Depends(current_user)):
    await require_perm(user, "ombor_korish")
    return await _materials_enriched()


@app.get("/api/alerts")
async def alerts(user=Depends(current_user)):
    await require_perm(user, "ombor_korish")
    return [m for m in await _materials_enriched() if m["kam"]]


@app.get("/api/finished-goods")
async def finished_goods(product_id: int | None = None, user=Depends(current_user)):
    await require_perm(user, "tayyor_mahsulot_korish")
    if product_id is not None:
        return await db.get_finished_goods(product_id)
    return await db.get_all_finished_goods()


# ═══ ISHLAB CHIQARISH ═══
@app.get("/api/production/templates")
async def production_templates(product_id: int, user=Depends(current_user)):
    await require_perm(user, "ishlab_chiqarish_korish")
    sh = await db.get_shablonlar(product_id, faqat_faol=True)
    return [{"id": s["id"], "kod": s.get("kod"), "nomi": s["nomi"],
             "chiqim": [{"nomi": c.get("blok_nomi") or c.get("block_kod"),
                         "soni": c["soni"]} for c in s.get("chiqim", [])]}
            for s in sh]


class ProductionIn(BaseModel):
    product_id: int
    entries: dict


@app.post("/api/production")
async def production_add(data: ProductionIn, user=Depends(current_user)):
    await require_perm(user, "ishlab_chiqarish_kiritish")
    ok, payload = await db.add_production(data.product_id, data.entries, user["id"])
    return {"ok": ok, "natija": payload} if ok else {"ok": False, "sabab": payload}


class UndoIn(BaseModel):
    product_id: int


@app.post("/api/production/undo")
async def production_undo(data: UndoIn, user=Depends(current_user)):
    await require_perm(user, "ishlab_chiqarish_kiritish")
    ok, xabar = await db.delete_last_production_with_restore(data.product_id)
    return {"ok": ok, "xabar": xabar}


@app.get("/api/production/today")
async def production_today(product_id: int, user=Depends(current_user)):
    await require_perm(user, "ishlab_chiqarish_korish")
    return await db.get_production_today(product_id)


# ═══ SOTUV ═══
class SalesIn(BaseModel):
    product_id: int
    block_kod: str
    miqdor: int


@app.post("/api/sales")
async def sales_add(data: SalesIn, user=Depends(current_user)):
    await require_perm(user, "sotuv_kiritish")
    if data.miqdor <= 0:
        raise HTTPException(400, "Miqdor 0 dan katta bo'lishi kerak")
    ok, xabar = await db.add_sales_log(data.product_id, data.block_kod,
                                       data.miqdor, user["id"])
    return {"ok": ok, "xabar": xabar}


@app.post("/api/sales/undo")
async def sales_undo(data: UndoIn, user=Depends(current_user)):
    await require_perm(user, "sotuv_kiritish")
    ok = await db.delete_last_sale(data.product_id)
    return {"ok": ok, "xabar": "Oxirgi sotuv bekor qilindi" if ok else "Bekor qilinadigan sotuv yo'q"}


@app.get("/api/sales/today")
async def sales_today(product_id: int, user=Depends(current_user)):
    await require_perm(user, "sotuv_korish")
    return await db.get_sales_today(product_id)


# ═══ OMBOR (materiallar) ═══
class MaterialIn(BaseModel):
    nomi: str
    qoldiq: float
    birlik: str


@app.post("/api/materials")
async def material_add(data: MaterialIn, user=Depends(current_user)):
    await require_perm(user, "ombor_kiritish")
    await db.add_material(data.nomi.strip(), data.qoldiq, data.birlik.strip())
    return {"ok": True}


class KirimIn(BaseModel):
    miqdor: float


@app.post("/api/materials/{material_id}/kirim")
async def material_kirim(material_id: int, data: KirimIn, user=Depends(current_user)):
    await require_perm(user, "ombor_kiritish")
    rows = await db.get_materials()
    hozir = next((r for r in rows if r[0] == material_id), None)
    if not hozir:
        raise HTTPException(404, "Material topilmadi")
    asl = hozir[4]
    qoshiladigan_baza, _ = birlikni_asosiyga(data.miqdor, asl)
    yangi_baza = max(0, hozir[2] + qoshiladigan_baza)
    await db.update_material_qoldiq(material_id, yangi_baza)
    return {"ok": True, "yangi_qoldiq": round(asosiydan_birlikga(yangi_baza, asl), 2),
            "birlik": asl}


class MinIn(BaseModel):
    min_chegara: float  # asl birlikda


@app.post("/api/materials/{material_id}/min")
async def material_min(material_id: int, data: MinIn, user=Depends(current_user)):
    await require_perm(user, "ombor_kiritish")
    rows = await db.get_materials()
    hozir = next((r for r in rows if r[0] == material_id), None)
    if not hozir:
        raise HTTPException(404, "Material topilmadi")
    baza, _ = birlikni_asosiyga(data.min_chegara, hozir[4])
    await db.set_min_chegara(material_id, baza)
    return {"ok": True}


class NarxIn(BaseModel):
    narx: float  # UZS, asl birlik uchun


@app.post("/api/materials/{material_id}/narx")
async def material_narx(material_id: int, data: NarxIn, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    # narx asl birlik uchun -> baza birlik uchun narxga aylantiramiz
    rows = await db.get_materials()
    hozir = next((r for r in rows if r[0] == material_id), None)
    if not hozir:
        raise HTTPException(404, "Material topilmadi")
    asl = hozir[4]
    # 1 asl birlik = koeff baza birlik; narx_baza = narx_asl / koeff
    koeff, _ = birlikni_asosiyga(1.0, asl)
    narx_baza = data.narx / koeff if koeff else data.narx
    await db.set_material_narx(material_id, narx_baza)
    return {"ok": True}


# ═══ INVENTARIZATSIYA ═══
class InventoryIn(BaseModel):
    product_id: int
    block_kod: str
    real_hisob: int
    izoh: str = ""


@app.post("/api/inventory")
async def inventory_add(data: InventoryIn, user=Depends(current_user)):
    await require_perm(user, "inventarizatsiya")
    fg = await db.get_finished_goods(data.product_id)
    blk = next((b for b in fg if b["kod"] == data.block_kod), None)
    bot_hisob = blk["qoldiq"] if blk else 0
    farq = await db.add_inventarizatsiya(
        data.product_id, bugungi_sana(), data.block_kod, bot_hisob,
        data.real_hisob, data.izoh, user["id"])
    return {"ok": True, "bot_hisob": bot_hisob, "real_hisob": data.real_hisob, "farq": farq}


# ═══ HISOBOTLAR ═══
@app.get("/api/reports/summary")
async def reports_summary(
    frm: str = Query(..., alias="from"),
    to: str = Query(...),
    product_id: int | None = None,
    user=Depends(current_user),
):
    await require_perm(user, "hisobot_korish")
    prod_qolip = await db.get_production_qolip_range(frm, to, product_id)
    prod_blocks = await db.get_production_blocks_range(frm, to, product_id)
    sales_blocks = await db.get_sales_blocks_range(frm, to, product_id)
    material_sarfi = await db.get_material_sarfi(frm, to, product_id)
    by_user_prod = await db.get_production_by_user_range(frm, to)
    by_user_sales = await db.get_sales_by_user_range(frm, to)
    jami_qolip = sum(v["qolip"] for v in prod_qolip.values())
    jami_sotuv = sum(b["qty"] for b in sales_blocks)
    daromad = sum(b["rev"] for b in sales_blocks)
    material_xarajat = sum(m["jami"] * m["narx"] for m in material_sarfi)
    return {
        "jami_qolip": jami_qolip, "jami_sotuv": jami_sotuv,
        "daromad": daromad, "material_xarajat": material_xarajat,
        "prod_blocks": prod_blocks, "sales_blocks": sales_blocks,
        "material_sarfi": material_sarfi,
        "by_user_prod": by_user_prod, "by_user_sales": by_user_sales,
    }


@app.get("/api/finance/summary")
async def finance_summary(
    frm: str = Query(..., alias="from"),
    to: str = Query(...),
    product_id: int | None = None,
    user=Depends(current_user),
):
    await require_perm(user, "moliya_korish")
    sales_blocks = await db.get_sales_blocks_range(frm, to, product_id)
    tannarx_map = await db.get_block_tannarx_map()
    daromad = sum(b["rev"] for b in sales_blocks)
    cogs = sum(b["qty"] * tannarx_map.get((b["product_id"], b["kod"]), 0.0)
               for b in sales_blocks)
    ombor_qiymati = await db.ombor_xom_qiymati()
    return {"daromad": daromad, "cogs": cogs, "foyda": daromad - cogs,
            "ombor_qiymati": ombor_qiymati}


@app.get("/api/tannarx")
async def tannarx(product_id: int, user=Depends(current_user)):
    await require_perm(user, "moliya_korish")
    return await db.tannarx_hisobla(product_id)


@app.get("/api/reports/excel")
async def reports_excel(
    frm: str = Query(..., alias="from"),
    to: str = Query(...),
    product_id: int | None = None,
    user=Depends(current_user),
):
    await require_perm(user, "excel_hisobot")
    import openpyxl
    from openpyxl.styles import Font

    prod_blocks = await db.get_production_blocks_range(frm, to, product_id)
    sales_blocks = await db.get_sales_blocks_range(frm, to, product_id)
    material_sarfi = await db.get_material_sarfi(frm, to, product_id)
    by_user_prod = await db.get_production_by_user_range(frm, to)
    by_user_sales = await db.get_sales_by_user_range(frm, to)

    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    def sheet(title, headers, rows):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.font = bold
        for r in rows:
            ws.append(r)

    wb.remove(wb.active)
    sheet("Ishlab chiqarish", ["Mahsulot", "Blok", "Soni"],
          [[b["product_nomi"], b["blok_nomi"], b["soni"]] for b in prod_blocks])
    sheet("Sotuv", ["Mahsulot", "Blok", "Soni", "Daromad (UZS)"],
          [[b["product_nomi"], b["blok_nomi"], b["qty"], b["rev"]] for b in sales_blocks])
    sheet("Material sarfi", ["Material", "Sarf", "Birlik", "Narx", "Summa"],
          [[m["nomi"], m["jami"], m["birlik"], m["narx"], m["jami"] * m["narx"]]
           for m in material_sarfi])
    sheet("Xodimlar (ishlab ch.)", ["Xodim", "Rol", "Qolip", "Ish haqi (UZS)"],
          [[u["ism"], u["rol"], u["qolip"], u["haq"]] for u in by_user_prod])
    sheet("Xodimlar (sotuv)", ["Xodim", "Rol", "Sotuv", "Daromad (UZS)"],
          [[u["ism"], u["rol"], u["qty"], u["rev"]] for u in by_user_sales])

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    fn = f"hisobot_{frm}_{to}.xlsx"
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


# ═══ SOZLAMA: mahsulot / blok / shablon / formula ═══
class ProductIn(BaseModel):
    kod: str
    nomi: str
    emoji: str = "📦"


@app.post("/api/products")
async def product_create(data: ProductIn, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    pid = await db.add_mahsulot(data.kod.strip(), data.nomi.strip(), data.emoji or "📦")
    return {"ok": True, "id": pid}


class ProductPatch(BaseModel):
    nomi: str | None = None
    emoji: str | None = None
    ishchi_haqi: float | None = None
    qoshimcha: float | None = None


@app.patch("/api/products/{product_id}")
async def product_patch(product_id: int, data: ProductPatch, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    cur = await db.get_mahsulot(product_id)
    if not cur:
        raise HTTPException(404, "Mahsulot topilmadi")
    if data.nomi is not None or data.emoji is not None:
        await db.update_mahsulot(product_id, data.nomi if data.nomi is not None else cur["nomi"],
                                 data.emoji if data.emoji is not None else cur.get("emoji"))
    if data.ishchi_haqi is not None:
        await db.set_mahsulot_ishchi_haqi(product_id, data.ishchi_haqi)
    if data.qoshimcha is not None:
        await db.set_mahsulot_qoshimcha(product_id, data.qoshimcha)
    return {"ok": True}


@app.get("/api/products/{product_id}/blocks")
async def product_blocks(product_id: int, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    return await db.get_bloklar(product_id)


class BlockIn(BaseModel):
    kod: str
    nomi: str
    olcham: str = ""
    qolip_dona: float = 0
    sotuv_narx: float = 0


@app.post("/api/products/{product_id}/blocks")
async def block_create(product_id: int, data: BlockIn, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    bid = await db.add_blok(product_id, data.kod.strip(), data.nomi.strip(),
                            data.olcham, data.qolip_dona, data.sotuv_narx)
    return {"ok": True, "id": bid}


class BlockPatch(BaseModel):
    nomi: str
    olcham: str = ""
    qolip_dona: float = 0
    sotuv_narx: float = 0


@app.patch("/api/blocks/{block_id}")
async def block_patch(block_id: int, data: BlockPatch, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    await db.update_blok(block_id, data.nomi.strip(), data.olcham, data.qolip_dona)
    await db.set_blok_sotuv_narx(block_id, data.sotuv_narx)
    return {"ok": True}


@app.delete("/api/blocks/{block_id}")
async def block_delete(block_id: int, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    await db.delete_blok(block_id)
    return {"ok": True}


@app.get("/api/products/{product_id}/templates-config")
async def templates_config(product_id: int, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    sh = await db.get_shablonlar(product_id, faqat_faol=False)
    return [{"id": s["id"], "kod": s.get("kod"), "nomi": s["nomi"],
             "chiqim": [{"block_kod": c["block_kod"], "soni": c["soni"],
                         "nomi": c.get("blok_nomi") or c["block_kod"]}
                        for c in s.get("chiqim", [])]} for s in sh]


class TemplateIn(BaseModel):
    kod: str
    nomi: str


@app.post("/api/products/{product_id}/templates")
async def template_create(product_id: int, data: TemplateIn, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    sid = await db.add_shablon(product_id, data.kod.strip(), data.nomi.strip())
    return {"ok": True, "id": sid}


class ChiqimIn(BaseModel):
    chiqim: list  # [{block_kod, soni}]


@app.post("/api/templates/{template_id}/chiqim")
async def template_chiqim(template_id: int, data: ChiqimIn, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    pairs = [(c["block_kod"], int(c["soni"])) for c in data.chiqim if int(c.get("soni", 0)) > 0]
    await db.set_shablon_chiqim(template_id, pairs)
    return {"ok": True}


@app.delete("/api/templates/{template_id}")
async def template_delete(template_id: int, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    await db.delete_shablon(template_id)
    return {"ok": True}


@app.get("/api/products/{product_id}/formula")
async def product_formula(product_id: int, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    # tuple: (nomi, miqdor, birlik, qoldiq_baza, birlik_baza, material_id, miqdor_asosiy, asl_birlik)
    rows = await db.get_qolip_formula(product_id)
    return [{"material_id": r[5], "nomi": r[0], "miqdor": r[1], "birlik": r[2],
             "asl_birlik": r[7]} for r in rows]


class FormulaIn(BaseModel):
    material_id: int
    miqdor: float
    birlik: str


@app.post("/api/products/{product_id}/formula")
async def formula_set(product_id: int, data: FormulaIn, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    await db.set_qolip_formula_item(product_id, data.material_id, data.miqdor, data.birlik.strip())
    return {"ok": True}


@app.delete("/api/products/{product_id}/formula/{material_id}")
async def formula_remove(product_id: int, material_id: int, user=Depends(current_user)):
    await require_perm(user, "sozlama_boshqaruv")
    await db.remove_qolip_formula_item(product_id, material_id)
    return {"ok": True}


# ═══ XODIMLAR ═══
ROLLAR = ["superadmin", "direktor", "omborchi", "ishchi", "sotuvchi", "hisobchi"]


@app.get("/api/users")
async def users_list(user=Depends(current_user)):
    await require_perm(user, "foydalanuvchi_boshqaruv")
    rows = await db.get_all_users()
    return [{"id": r["id"], "ism": r.get("ism"), "rol": r.get("rol"),
             "telefon": r.get("telefon"), "faol": r.get("faol"),
             "parol_bor": bool(r.get("parol_hash"))} for r in rows]


async def _set_credentials(conn, user_id, telefon, parol):
    telefon = (telefon or "").strip() or None
    parol_hash = auth.hash_password(parol) if parol else None
    if telefon is not None and parol_hash is not None:
        await conn.execute("UPDATE users SET telefon=$1, parol_hash=$2 WHERE id=$3",
                           telefon, parol_hash, user_id)
    elif telefon is not None:
        await conn.execute("UPDATE users SET telefon=$1 WHERE id=$2", telefon, user_id)
    elif parol_hash is not None:
        await conn.execute("UPDATE users SET parol_hash=$1 WHERE id=$2", parol_hash, user_id)


class UserCreateIn(BaseModel):
    ism: str
    telefon: str
    parol: str
    rol: str


@app.post("/api/users")
async def user_create(data: UserCreateIn, user=Depends(current_user)):
    await require_perm(user, "foydalanuvchi_boshqaruv")
    if data.rol not in ROLLAR:
        raise HTTPException(400, "Noto'g'ri rol")
    new_id = -abs(int(time.time() * 1000) % 1_000_000_000) or -1
    pool = await get_pool()
    try:
        async with pool.acquire() as conn:
            async with conn.transaction():
                await conn.execute(
                    "INSERT INTO users (id, ism, username, rol, faol) VALUES ($1,$2,$3,$4,TRUE)",
                    new_id, data.ism.strip(), "", data.rol)
                await _set_credentials(conn, new_id, data.telefon, data.parol)
    except Exception as e:
        if "telefon" in str(e).lower() or "unique" in str(e).lower():
            raise HTTPException(400, "Bu telefon raqami allaqachon band")
        raise HTTPException(400, f"Xatolik: {e}")
    return {"ok": True, "id": new_id}


class CredsIn(BaseModel):
    telefon: str | None = None
    parol: str | None = None


@app.post("/api/users/{user_id}/credentials")
async def user_credentials(user_id: int, data: CredsIn, user=Depends(current_user)):
    await require_perm(user, "foydalanuvchi_boshqaruv")
    pool = await get_pool()
    try:
        async with pool.acquire() as conn:
            await _set_credentials(conn, user_id, data.telefon, data.parol)
    except Exception as e:
        if "telefon" in str(e).lower() or "unique" in str(e).lower():
            raise HTTPException(400, "Bu telefon raqami allaqachon band")
        raise HTTPException(400, f"Xatolik: {e}")
    return {"ok": True}


class RolIn(BaseModel):
    rol: str


@app.patch("/api/users/{user_id}")
async def user_update_rol(user_id: int, data: RolIn, user=Depends(current_user)):
    await require_perm(user, "foydalanuvchi_boshqaruv")
    if data.rol not in ROLLAR:
        raise HTTPException(400, "Noto'g'ri rol")
    await db.update_user_rol(user_id, data.rol)
    return {"ok": True}


@app.delete("/api/users/{user_id}")
async def user_deactivate(user_id: int, user=Depends(current_user)):
    await require_perm(user, "foydalanuvchi_boshqaruv")
    if user_id == user["id"]:
        raise HTTPException(400, "O'z hisobingizni bloklay olmaysiz")
    await db.delete_user(user_id)
    return {"ok": True}


@app.get("/api/health")
async def health():
    return {"ok": True}


# ── Mobil web ilova (static) ──
_static_dir = os.path.join(os.path.dirname(__file__), "static")
if os.path.isdir(_static_dir):
    app.mount("/", StaticFiles(directory=_static_dir, html=True), name="static")
